from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import CatalogBundle


XLSX_HEADER_MAP = {
    "ID": "id",
    "Nombre Oficial": "nombre",
    "Sigla": "sigla",
    "Categoria": "categoria",
    "Categoría": "categoria",
    "Sector": "sector",
    "Dependencia": "dependencia",
    "Region": "region",
    "Región": "region",
    "Sitio Web Oficial": "sitio_web",
    "URL Portal de Empleos": "url_empleo",
    "Plataforma Empleo": "plataforma_empleo",
    "En empleospublicos.cl": "publica_en_empleospublicos",
    "Dificultad Scraping": "dificultad_scraping",
    "Requiere JS": "requiere_js",
    "Notas Tecnicas": "notas_tecnicas",
    "Notas Técnicas": "notas_tecnicas",
    "Poblacion Censo 2024": "poblacion_censo2024",
    "Población Censo 2024": "poblacion_censo2024",
    "Estado Verificacion": "estado_verificacion",
    "Estado Verificación": "estado_verificacion",
}


class CatalogLoader:
    def __init__(self, json_path: str | Path | None = None, xlsx_path: str | Path | None = None) -> None:
        root = Path(__file__).resolve().parents[2]
        self.json_path = Path(json_path) if json_path else root / "repositorio_instituciones_publicas_chile.json"
        self.xlsx_path = Path(xlsx_path) if xlsx_path else root / "repositorio_instituciones_publicas_chile_v3.xlsx"

    def load(self, *, prefer_json: bool = True) -> CatalogBundle:
        json_bundle = self.load_json() if self.json_path.exists() else None
        xlsx_bundle = self.load_xlsx() if self.xlsx_path.exists() else None
        if prefer_json and json_bundle:
            if xlsx_bundle:
                return self._merge(json_bundle, xlsx_bundle)
            return json_bundle
        if xlsx_bundle:
            return xlsx_bundle if not json_bundle else self._merge(xlsx_bundle, json_bundle)
        if json_bundle:
            return json_bundle
        raise FileNotFoundError("No se encontro ningun catalogo JSON/XLSX disponible.")

    def load_json(self) -> CatalogBundle:
        payload = json.loads(self.json_path.read_text(encoding="utf-8-sig"))
        instituciones = payload.get("instituciones", payload if isinstance(payload, list) else [])
        metadata = payload.get("metadata", {}) if isinstance(payload, dict) else {}
        return CatalogBundle(instituciones=[self._normalize(item) for item in instituciones], metadata=metadata, source="json")

    def load_xlsx(self) -> CatalogBundle:
        workbook = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=4, max_row=4))]
            records: list[dict[str, Any]] = []
            for row in sheet.iter_rows(min_row=5, values_only=True):
                if not any(row):
                    continue
                raw = {XLSX_HEADER_MAP.get(str(header).strip(), str(header).strip()): value for header, value in zip(headers, row)}
                records.append(self._normalize(raw))
            empty_status = all(not item.get("estado_verificacion") for item in records)
            return CatalogBundle(
                instituciones=records,
                metadata={"estado_verificacion_vacio": empty_status, "sheet": sheet.title},
                source="xlsx",
            )
        finally:
            workbook.close()

    def _merge(self, primary: CatalogBundle, secondary: CatalogBundle) -> CatalogBundle:
        by_id = {item.get("id"): dict(item) for item in primary.instituciones}
        for item in secondary.instituciones:
            ident = item.get("id")
            if ident not in by_id:
                by_id[ident] = dict(item)
                continue
            for key, value in item.items():
                if by_id[ident].get(key) in (None, "", []):
                    by_id[ident][key] = value
        metadata = dict(primary.metadata)
        metadata["secondary_source"] = secondary.source
        metadata["secondary_metadata"] = secondary.metadata
        return CatalogBundle(instituciones=list(by_id.values()), metadata=metadata, source=primary.source)

    @staticmethod
    def _normalize(record: dict[str, Any]) -> dict[str, Any]:
        normalized = {str(key): value for key, value in record.items()}
        if normalized.get("id") not in (None, ""):
            try:
                normalized["id"] = int(normalized["id"])
            except (TypeError, ValueError):
                pass
        return normalized
